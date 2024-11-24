from openpyxl import Workbook

#creates dictionary to reference
week_temperatures = {
    'Monday': 54,
    'Tuesday': 60,
    'Wednesday': 62,
    'Thursday': 57,
    'Friday': 71,
}

#creates workbook and uses first sheet in excell
workbook = Workbook()
worksheet = workbook.active

#titles for excell file
worksheet.title = 'Daily Temperatures'
worksheet.cell(1,1, 'Day')
worksheet.cell(1,2, 'Temperature(F)')

row_index = 2 #starts on row 2 since titles are in row 1

#for loop to go through dictionary and place data in cells
for day, temp in week_temperatures.items():
    worksheet.cell(row_index, 1, day) #day key from above dictionary in row 1
    worksheet.cell(row_index, 2, temp) #temp value from dictionary in row 2
    row_index += 1 #adds 1 to row index after each iteration to go to next row

workbook.save('temperatures.xlsx') #saves workbook