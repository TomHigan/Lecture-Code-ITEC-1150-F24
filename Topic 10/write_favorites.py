from openpyxl import Workbook

#create lists
favorite_foods = ['Pizza', 'Cake', 'Broccoli']
color_favorites = ['Blue', 'Purple', 'Red']

workbook = Workbook() #defines the workbook variable with the built in workbook container

worksheet = workbook.active #moves to the first sheet

worksheet.title = 'Favorite Things' #titles the excell worksheet

worksheet.cell(1,1, 'Favorite Foods') #puts Favorite Foods in cell A1

for index, food in enumerate(favorite_foods):
    worksheet.cell(index+2, 1, food) #Places list items in cells. index +2 since it starts at 0 and A1 has a title, and column 1.

worksheet.cell(1,2, 'Favorite Colors') #puts Favorite Colors in cell B1

for index, colors in enumerate(color_favorites):
    worksheet.cell(index+2,2, colors) #Places list items in cells. index +2 since it starts at 0 and A1 has a title, and column 2.

workbook.save('favorites.xlsx') #saves workbook
